import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ========= FILE =========
file_path = r"C:\Users\VSOLUTIONS\Documents\DATA.xlsx"

wb = openpyxl.load_workbook(file_path)
sheet_links = wb["Sheet1"]

# ========= CREATE OUTPUT SHEET =========
if "Sheet2" in wb.sheetnames:
    sheet_output = wb["Sheet2"]
else:
    sheet_output = wb.create_sheet("Sheet2")

# ========= HEADERS =========
headers = [
    "No",
    "Company Link",
    "Address",
    "Phone",
    "Email",
    "Website",
    "LinkedIn"
]

# Write headers only once
if sheet_output.max_row == 1 and sheet_output["A1"].value is None:
    for col, header in enumerate(headers, start=1):
        sheet_output.cell(row=1, column=col).value = header

# ========= CHROME =========
options = webdriver.ChromeOptions()
options.add_argument(r"--user-data-dir=C:\selenium-profile")
options.add_argument("--start-maximized")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 15)

output_row = sheet_output.max_row + 1
row = 2
link_number = 1

# ========= START LOOP =========
while sheet_links[f"A{row}"].value:

    link = sheet_links[f"A{row}"].value

    try:
        driver.get(link)

        # wait page load
        wait.until(
            EC.presence_of_element_located(
                (By.CLASS_NAME, "exhibitors-details__info")
            )
        )

        time.sleep(2)

        # ========= DEFAULT VALUES =========
        address = ""
        phone = ""
        email = ""
        website = ""
        linkedin = ""

        # ========= ADDRESS =========
        try:
            address = driver.find_element(
                By.XPATH,
                "//i[contains(@class,'icon-map-marker')]/parent::span/strong"
            ).text.strip()
        except:
            pass

        # ========= PHONE =========
        try:
            phone = driver.find_element(
                By.XPATH,
                "//i[contains(@class,'icon-phone')]/parent::a/strong"
            ).text.strip()
        except:
            pass

        # ========= EMAIL =========
        try:
            email = driver.find_element(
                By.XPATH,
                "//a[contains(@href,'mailto:')]/strong"
            ).text.strip()
        except:
            pass

        # ========= WEBSITE =========
        try:
            website = driver.find_element(
                By.XPATH,
                "//a[contains(@href,'http') and .//i[contains(@class,'icon-world')]]"
            ).get_attribute("href")
        except:
            pass

        # ========= LINKEDIN =========
        try:
            linkedin = driver.find_element(
                By.XPATH,
                "//a[contains(@href,'linkedin.com')]"
            ).get_attribute("href")
        except:
            pass

        # ========= SAVE DATA =========
        sheet_output.cell(row=output_row, column=1).value = link_number
        sheet_output.cell(row=output_row, column=2).value = link
        sheet_output.cell(row=output_row, column=3).value = address
        sheet_output.cell(row=output_row, column=4).value = phone
        sheet_output.cell(row=output_row, column=5).value = email
        sheet_output.cell(row=output_row, column=6).value = website
        sheet_output.cell(row=output_row, column=7).value = linkedin

        print(f"{link_number} DONE")

        output_row += 1

        # SAVE every 10 links
        if link_number % 10 == 0:
            wb.save(file_path)

    except Exception as e:
        print(f"{link_number} ERROR : {e}")

    row += 1
    link_number += 1

# ========= FINAL SAVE =========
wb.save(file_path)

driver.quit()

print("🔥 ALL DONE")
